from openpyxl import load_workbook
import json
import os
from datetime import datetime

# ================= CONFIG =================
HEADER_ROW = 22
DATA_START_ROW = 23
# =========================================


# ================= HELPER =================
def col(letter):
    """Convert Excel column letter to index (0-based)"""
    num = 0
    for c in letter:
        num = num * 26 + (ord(c.upper()) - ord("A") + 1)
    return num - 1


def val(row, letter):
    v = row[col(letter)]
    if v in ("", None, "null", "NULL"):
        return None
    return v


def as_str(v):
    if v is None:
        return ""
    if isinstance(v, float):
        return str(int(v))
    return str(v)


def as_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def as_bool(v):
    if isinstance(v, bool):
        return v
    if v in (1, "1", "Y", "y", "true", "TRUE"):
        return True
    return False


def as_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat() + "Z"
    return str(v)
# =========================================


# ================= MAIN =================
def generate_payload(
    excel_path,
    output_dir="payloads",
    debug=False
):
    os.makedirs(output_dir, exist_ok=True)

    wb = load_workbook(excel_path, data_only=True)
    generated = []

    # ---------- PROCESS BOTH SHEETS ----------
    for sheet_name in ["API_Doc_Agent_Broker", "API_Doc_Company"]:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            if not any(row):
                continue

            # ===== FILE NAME (A–D) =====
            file_name = (
                f"{as_str(val(row,'A'))}_"
                f"{as_str(val(row,'B'))}_"
                f"{as_str(val(row,'C'))}_"
                f"{as_str(val(row,'D'))}.json"
            )
            file_path = os.path.join(output_dir, file_name)

            # ===== COMMON PAYLOAD =====
            payload = {
                "system_code": as_str(val(row, "E")),
                "approval_flag": as_bool(val(row, "F")),
                "call_back_url": as_str(val(row, "G")),
                "approval": {
                    "doc_no": as_str(val(row, "H")),
                    "doc_ref": as_str(val(row, "I")),
                    "net_amount": as_float(val(row, "J")),
                    "remark": as_str(val(row, "K")),
                    "doc_meta_data": {
                        "group_code": as_str(val(row, "L")),
                        "account_code": as_str(val(row, "M")),
                        "disbursement_type": as_str(val(row, "N")),
                        "vat": as_float(val(row, "O")),
                        "wht": as_float(val(row, "P")),
                        "amount": as_float(val(row, "Q")),
                    },
                    "payees": {
                        "ga_no": val(row, "R"),
                        "payee_info": [],
                        "committees": []
                    }
                }
            }

            # ===== AGENT / BROKER =====
            if sheet_name == "API_Doc_Agent_Broker":
                payload["approval"]["payees"]["payee_info"].append({
                    "payment_type": as_str(val(row, "S")),
                    "payee_type": as_str(val(row, "T")),
                    "title": {
                        "en_US": as_str(val(row, "U")),
                        "th_TH": as_str(val(row, "V")),
                    },
                    "first_name": {
                        "en_US": as_str(val(row, "W")),
                        "th_TH": as_str(val(row, "X")),
                    },
                    "last_name": {
                        "en_US": as_str(val(row, "Y")),
                        "th_TH": as_str(val(row, "Z")),
                    },
                    "mobile_no": as_str(val(row, "AA")),
                    "identity_type": as_str(val(row, "AB")),
                    "identity_no": as_str(val(row, "AC")),
                    "code": as_str(val(row, "AD")),
                    "bank_info": {
                        "bank_code": as_str(val(row, "AE")),
                        "branch_code": as_str(val(row, "AF")),
                        "account_no": as_str(val(row, "AG")),
                        "account_name": as_str(val(row, "AH")),
                        "media_clearing_type_code": as_str(val(row, "AI")),
                    },
                    "cheque_info": val(row, "AJ"),
                    "k_trade_info": val(row, "AK"),
                    "relation": as_str(val(row, "AL")),
                    "date_of_birth": as_date(val(row, "AM")),
                    "pay_rate": as_float(val(row, "AN")),
                    "amount": as_float(val(row, "AO")),
                    "delivery_address_1": as_str(val(row, "AP")),
                    "delivery_address_2": as_str(val(row, "AQ")),
                    "mailing_zip_code": as_str(val(row, "AR")),
                    "tax": val(row, "AS")
                })

                payload["approval"]["payees"]["committees"] = []

            # ===== COMPANY =====
            else:
                payload["approval"]["payees"]["payee_info"].append({
                    "payment_type": as_str(val(row, "S")),
                    "payee_type": as_str(val(row, "T")),
                    "payee_name": {
                        "en_US": as_str(val(row, "U")),
                        "th_TH": as_str(val(row, "V")),
                    },
                    "title": {
                        "en_US": as_str(val(row, "W")),
                        "th_TH": as_str(val(row, "X")),
                    },
                    "mobile_no": as_str(val(row, "Y")),
                    "identity_type": as_str(val(row, "Z")),
                    "identity_no": as_str(val(row, "AA")),
                    "code": as_str(val(row, "AB")),
                    "bank_info": {
                        "bank_code": as_str(val(row, "AC")),
                        "branch_code": as_str(val(row, "AD")),
                        "account_no": as_str(val(row, "AE")),
                        "account_name": as_str(val(row, "AF")),
                        "media_clearing_type_code": as_str(val(row, "AG")),
                    },
                    "cheque_info": val(row, "AH"),
                    "k_trade_info": val(row, "AI"),
                    "relation": as_str(val(row, "AJ")),
                    "date_of_birth": as_date(val(row, "AK")),
                    "pay_rate": as_float(val(row, "AL")),
                    "amount": as_float(val(row, "AM")),
                    "delivery_address_1": as_str(val(row, "AN")),
                    "delivery_address_2": as_str(val(row, "AO")),
                    "mailing_zip_code": as_str(val(row, "AP")),
                    "tax": val(row, "AQ")
                })

                payload["approval"]["payees"]["committees"].append({
                    "title": {
                        "en_US": as_str(val(row, "AR")),
                        "th_TH": as_str(val(row, "AS")),
                    },
                    "first_name": {
                        "en_US": as_str(val(row, "AT")),
                        "th_TH": as_str(val(row, "AU")),
                    },
                    "last_name": {
                        "en_US": as_str(val(row, "AV")),
                        "th_TH": as_str(val(row, "AW")),
                    },
                    "identity_type": as_str(val(row, "AX")),
                    "identity_no": as_str(val(row, "AY")),
                })

            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)

            generated.append(file_path)

            if debug:
                print(f"Generated: {file_path}")

    return generated
# =========================================
